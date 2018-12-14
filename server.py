from flask import Flask
from flask import render_template
from flask import request, redirect, abort, url_for
from flask import send_file

from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl import load_workbook

from EagleEye import * 
    
import settings

een = EagleEye() 


app = Flask(__name__)

@app.route("/", methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_data():

    esn = request.form.get('esn')
    start_timestamp = request.form.get('start_time')
    end_timestamp = request.form.get('end_time')

    een.login(username=settings.username, password=settings.password)

    camera = een.find_by_esn(esn)

    camera.get_video_list(instance=een, start_timestamp=start_timestamp, end_timestamp=end_timestamp)

    wb = Workbook()
    ws = wb.active

    for video in camera.videos:
        ws.append([camera.camera_id, video[0], video[1]])
        #een._format_url_for_download(esn=camera.camera_id, start_time=video[0], end_time=video[1])

    filename = f"./uploads/{esn}-{start_timestamp}.xlsx"
 
    wb.save(filename)

    return send_file(filename, as_attachment=True)


@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)

        f = request.files['file']
        location = secure_filename(f.filename)
        f.save('./uploads/' + location)
        return redirect('/view/%s' % location, code=302)

@app.route('/view/<path:filename>', methods=['GET'])
def view(filename):
    wb = load_workbook('./uploads/' + filename)
    ret_obj = {}

    if wb:
        ws = wb.active

        for row in ws.iter_rows(row_offset=2):
            esn = row[0].value
            start = row[1].value
            end = row[2].value

            if esn:
                if esn in ret_obj:
                    ret_obj[esn].append(
                        {
                            'start_timestamp': start,
                            'end_timestamp': end
                        }   
                    )
                else:
                    ret_obj[esn] = [{
                        'start_timestamp': start,
                        'end_timestamp': end
                    }]

    return render_template('upload.html', template_values=ret_obj)
